"""
Preprocess newfactory.xlsx: openpyxl reads rich-text cells correctly but
SheetJS (xlsx 0.18) returns empty `v` for them. Rewrite each rich-text
cell as a plain string so the TS importer can consume it.
"""
import openpyxl
from openpyxl.cell.rich_text import CellRichText
import sys

PATH = r"C:\Users\Eli\cursor-projects\albadi\albadi-crm\newfactory.xlsx"
wb = openpyxl.load_workbook(PATH, rich_text=True)
total = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, CellRichText):
                cell.value = str(v)
                total += 1
            elif isinstance(v, str) and not v.strip():
                # leave alone
                pass
wb.save(PATH)
print(f"Flattened {total} rich-text cells -> plain strings.")
