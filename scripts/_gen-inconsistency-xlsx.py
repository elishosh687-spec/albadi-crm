#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the factory-inconsistency workbook (Hebrew + Chinese per row) for the factory meeting."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

HEAD = ["类别 קטגוריה", "尺寸/产品 מידה", "数量 כמות", "变体 וריאנט",
        "说明（中文）", "הסבר (עברית)", "记录 A", "记录 B"]

hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=13, color="1F3864")
warn_fill = PatternFill("solid", fgColor="FCE4D6")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="center")

def build_sheet(ws, title_zh, subtitle, rows):
    ws.sheet_view.rightToLeft = False
    ws.merge_cells("A1:H1")
    ws["A1"] = title_zh
    ws["A1"].font = title_font
    ws.merge_cells("A2:H2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="808080")
    for c, h in enumerate(HEAD, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = wrap; cell.border = border
    for i, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.alignment = wrap; cell.border = border
            if c in (7, 8):
                cell.fill = warn_fill; cell.font = Font(bold=True)
    widths = [20, 22, 12, 12, 30, 26, 20, 20]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = "A4"

# ─────────────────────────── 2D ───────────────────────────
rows_2d = [
    # p6 self-contradictions — price
    ["价格 מחיר", "H15×W20", "3000→5000", "无手挽 בלי ידית", "数量增加价格反而上涨", "יותר כמות אך יקר יותר", "3000: ¥0.53", "5000: ¥0.55"],
    ["价格 מחיר", "H15×W20", "5000", "无手挽 בלי ידית", "覆膜比不覆膜便宜（不合理）", "למינציה זולה מרגיל", "普通 ¥0.55", "覆膜 ¥0.46"],
    ["价格 מחיר", "H15×W20", "10000", "无手挽 בלי ידית", "覆膜比不覆膜便宜（不合理）", "למינציה זולה מרגיל", "普通 ¥0.37", "覆膜 ¥0.32"],
    ["价格 מחיר", "H15×W20", "5000", "手挽 ידית", "带手挽比不带手挽便宜（不合理）", "עם ידית זול מבלי", "带手挽 ¥0.41", "无手挽 ¥0.55"],
    ["价格 מחיר", "H15×W20", "10000", "手挽 ידית", "带手挽比不带手挽便宜（不合理）", "עם ידית זול מבלי", "带手挽 ¥0.33", "无手挽 ¥0.37"],
    # bigger-cheaper within 2D
    ["价格 מחיר", "H10×W15 → H15×W20", "3000", "带手挽 עם ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.78", "大 ¥0.58"],
    ["价格 מחיר", "H10×W15 → H15×W20", "3000", "无手挽 בלי ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.74", "大 ¥0.53"],
    ["价格 מחיר", "H10×W15 → H15×W20", "5000", "带手挽 עם ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.73", "大 ¥0.41"],
    ["价格 מחיר", "H25×W25 → H30×W40", "5000", "带手挽 עם ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.63", "大 ¥0.60"],
    ["价格 מחיר", "H10×W15 → H15×W20", "5000", "无手挽 בלי ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.71", "大 ¥0.55"],
    ["价格 מחיר", "H25×W25 → H30×W40", "5000", "无手挽 בלי ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.59", "大 ¥0.55"],
    ["价格 מחיר", "H10×W15 → H15×W20", "10000", "带手挽 עם ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.71", "大 ¥0.33"],
    ["价格 מחיר", "H25×W25 → H30×W40", "10000", "带手挽 עם ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.60", "大 ¥0.51"],
    ["价格 מחיר", "H10×W15 → H15×W20", "10000", "无手挽 בלי ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.68", "大 ¥0.37"],
    ["价格 מחיר", "H25×W25 → H30×W40", "10000", "无手挽 בלי ידית", "更大的袋子价格反而更低", "גדולה זולה מקטנה", "小 ¥0.56", "大 ¥0.46"],
    # real quote same spec
    ["价格 מחיר", "65×55 (报价)", "12000", "无手挽/普通", "同款袋子两个报价价格不同", "אותה שקית 2 מחירים", "¥0.90 (4TIMEAL4)", "¥0.91 (UWO0DZIC)"],
    # shipping
    ["运输 שילוח", "30×20 (PANLUIB8)", "3000", "—", "每箱数量导致隐含厚度约2.5毫米（80g袋不可能）", "עובי משתמע 2.5מ״מ", "300 个/箱", "≈2.5 mm/个"],
    ["运输 שילוח", "36×45 (FIRBM6CX)", "5000", "—", "隐含厚度约1.6毫米（偏高）", "עובי משתמע 1.6מ״מ", "250 个/箱", "≈1.6 mm/个"],
    ["运输 שילוח", "65×55 (报价)", "12000", "—", "同款袋子每箱数量相差一倍", "אותה שקית אריזה כפולה", "300 个/箱 (4TIMEAL4)", "150 个/箱 (UWO0DZIC)"],
]

# ─────────────────────────── 3D ───────────────────────────
rows_3d = [
    ["价格 מחיר", "H15×D5×W20 (p7)", "10000", "手挽 ידית", "带手挽与不带手挽价格相同（应更贵）", "עם ידית = בלי ידית", "带手挽 ¥0.40", "无手挽 ¥0.40"],
    ["运输 שילוח", "—", "—", "—", "未发现不一致 ✓", "לא נמצאו סתירות ✓", "—", "—"],
]

ws2 = wb.active; ws2.title = "二维袋 2D"
build_sheet(ws2, "二维袋 (2D) — 价格与运输不一致  |  דו-מימד — סתירות מחיר ושילוח",
            "袋子结构：无侧折 / 平口袋。数量 3000 以上。", rows_2d)
ws3 = wb.create_sheet("三维袋 3D")
build_sheet(ws3, "三维袋 (3D) — 价格与运输不一致  |  תלת-מימד — סתירות מחיר ושילוח",
            "袋子结构：有侧折。数量 3000 以上。（几乎没有不一致）", rows_3d)

out = "/Users/eli/Downloads/albadi-factory-inconsistencies.xlsx"
wb.save(out)
print("saved:", out)
print("2D rows:", len(rows_2d), "| 3D rows:", len(rows_3d))
