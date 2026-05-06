import os, requests, json, time
from datetime import datetime, timedelta

TOKEN = os.environ.get('MANYCHAT_TOKEN', '')
if not TOKEN:
    raise SystemExit("Set MANYCHAT_TOKEN in environment (see ../.env)")
H = {'Authorization': f'Bearer {TOKEN}', 'Content-Type': 'application/json'}
BASE = 'https://api.manychat.com/fb'
TIMEOUT = 10  # seconds per request

TODAY = datetime.now().date()

# Tag IDs and scores
TAG_SCORES = {
    84622722: ('בתהליך', 40),
    84604876: ('מעוניין', 30),
    84644793: ('הצעה_טלפון', 25),
    84604872: ('ליד_חדש', 20),
    84644778: ('הצעה_בוט', 15),
    84622721: ('לא_ענה', 5),
}

SKIP_TAGS = {84604877, 84604878}  # לא_רלוונטי, לקוח

# All subscriber IDs we know about
ALL_SUBS = [
    "1290975646", "335237336", "843866619", "1567115769", "2035644170",
    "1884294789", "1602697859", "933250256", "1945485008", "2121695200",
    "21902603", "342493590", "1342391971", "647013452", "235009133",
    "1109877399", "1233780185", "1168653412", "1745508158", "1559024601",
    "940287852", "969554152", "24594158", "1513055758", "1986772872",
    "3658499", "1890126495", "248319497", "221677737", "347894123",
    "869425808", "1768242677", "956589647", "771607363", "1720207271",
    "774945448", "1701651968", "1258938556", "306431271",
]

# Phone lookup from CSV
PHONES = {
    "1290975646": "972502040889", "335237336": "972054220443",
    "2035644170": "972558813450", "1884294789": "972546655546",
    "1602697859": "972543188816", "933250256": "972528903144",
    "2121695200": "972546610889", "21902603": "972527042481",
    "342493590": "972509000066", "1342391971": "972504759040",
    "647013452": "972549818844", "235009133": "972552548897",
    "1109877399": "972535201414", "1233780185": "972525163329",
    "1168653412": "972522424855", "1745508158": "972546261318",
    "940287852": "972525253040", "969554152": "972543990742",
    "24594158": "972506006656", "1513055758": "972508861985",
    "1986772872": "972539262485", "3658499": "972545221334",
    "1890126495": "972526188677", "248319497": "972544450094",
    "221677737": "972528792133", "347894123": "972505798761",
    "869425808": "972542133378", "1768242677": "972052666106",
    "1559024601": "972529041524",
}

# Field IDs
F_NOTES = 14447147
F_QUOTE_TOTAL = 14447148
F_FOLLOW_UP = 14445938
F_LAST_CONTACT = 14447151
F_LEAD_SCORE = 14445937
F_QUANTITY = 14356831

def get_field(fields, field_id):
    """Extract field value from subscriber's custom fields list."""
    for f in fields:
        if f.get('id') == field_id:
            return f.get('value', '')
    return ''

def quantity_score(qty):
    if qty >= 10000: return 30
    if qty >= 5000: return 20
    if qty >= 3000: return 15
    if qty >= 1000: return 10
    return 15  # unknown = potentially big

def urgency_score(follow_up_str):
    if not follow_up_str:
        return 10  # no date = treat as today
    try:
        fu_date = datetime.strptime(follow_up_str[:10], '%Y-%m-%d').date()
        diff = (TODAY - fu_date).days
        if diff >= 3: return 20
        if diff >= 1: return 15
        if diff == 0: return 10
        return 0
    except:
        return 10

def silence_score(last_contact_str):
    if not last_contact_str:
        return 10
    try:
        lc_date = datetime.strptime(last_contact_str[:10], '%Y-%m-%d').date()
        diff = (TODAY - lc_date).days
        if diff >= 7: return 10
        if diff >= 4: return 7
        if diff >= 2: return 3
        return 0
    except:
        return 5

def save_score_to_manychat(sid, score):
    """Save lead_score back to ManyChat."""
    try:
        r = requests.post(f'{BASE}/subscriber/setCustomFields', headers=H,
                          json={"subscriber_id": sid, "fields": [
                              {"field_id": F_LEAD_SCORE, "field_value": score}
                          ]}, timeout=TIMEOUT)
        return r.json().get('status') == 'success'
    except:
        return False

def save_to_excel(leads, all_leads_including_skipped):
    """Save all leads to Excel with formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("\n  openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'], stdout=subprocess.DEVNULL)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: Active Leads (sorted by score) ---
    ws = wb.active
    ws.title = "Active Leads"

    headers = ['#', 'Priority', 'Score', 'Name', 'Phone', 'Status', 'Quantity',
               'Quote (ILS)', 'Follow-up', 'Last Contact', 'Notes', 'Score Breakdown']

    # Header style
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Priority colors
    pri_fills = {
        'CALL NOW': PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid'),
        'HIGH': PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid'),
        'MEDIUM': PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),
        'LOW': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
    }

    for i, lead in enumerate(leads):
        row = i + 2
        score = lead['score']
        if score >= 70: pri = 'CALL NOW'
        elif score >= 50: pri = 'HIGH'
        elif score >= 35: pri = 'MEDIUM'
        else: pri = 'LOW'

        phone = lead['phone'] if lead['phone'] else ''
        qty = int(lead.get('quantity', 0)) if lead.get('quantity', 0) > 0 else ''
        qt = int(lead['quote_total']) if lead['quote_total'] > 0 else ''

        values = [
            i + 1, pri, score, lead['name'], phone, lead['tag'],
            qty, qt, lead['follow_up'], lead['last_contact'],
            lead['notes'], lead['breakdown'],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=(col == 10))  # wrap notes

        # Color priority cell
        ws.cell(row=row, column=2).fill = pri_fills.get(pri, PatternFill())
        if pri == 'CALL NOW':
            ws.cell(row=row, column=2).font = Font(color='FFFFFF', bold=True)

    # Column widths
    cols = ['A','B','C','D','E','F','G','H','I','J','K','L']
    widths = [5, 12, 7, 25, 15, 15, 10, 12, 12, 12, 50, 35]
    for col_letter, w in zip(cols, widths):
        ws.column_dimensions[col_letter].width = w

    # --- Sheet 2: All Leads (including inactive) ---
    ws2 = wb.create_sheet("All Leads")

    headers2 = ['Name', 'Phone', 'Status', 'Score', 'Quote (ILS)',
                'Follow-up', 'Last Contact', 'Notes', 'Active']

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, lead in enumerate(all_leads_including_skipped):
        row = i + 2
        phone = lead.get('phone', '')
        qt = int(lead['quote_total']) if lead.get('quote_total', 0) > 0 else ''
        active = 'Yes' if lead.get('active', True) else 'No'

        values = [
            lead['name'], phone, lead['tag'], lead.get('score', '-'),
            qt, lead.get('follow_up', ''), lead.get('last_contact', ''),
            lead.get('notes', ''), active,
        ]

        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border

        if not lead.get('active', True):
            gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            for col in range(1, len(values) + 1):
                ws2.cell(row=row, column=col).fill = gray

    cols2 = ['A','B','C','D','E','F','G','H','I']
    widths2 = [25, 15, 15, 7, 12, 12, 12, 50, 8]
    for col_letter, w in zip(cols2, widths2):
        ws2.column_dimensions[col_letter].width = w

    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet("Summary")
    ws3.cell(row=1, column=1, value="Daily Call Summary").font = Font(bold=True, size=14)
    ws3.cell(row=2, column=1, value=f"Date: {TODAY}")
    ws3.cell(row=4, column=1, value="Priority").font = Font(bold=True)
    ws3.cell(row=4, column=2, value="Count").font = Font(bold=True)

    call_now = len([l for l in leads if l['score'] >= 70])
    high = len([l for l in leads if 50 <= l['score'] < 70])
    medium = len([l for l in leads if 35 <= l['score'] < 50])
    low = len([l for l in leads if l['score'] < 35])

    summary_data = [
        ('CALL NOW', call_now, 'FF4444'),
        ('HIGH', high, 'FF8C00'),
        ('MEDIUM', medium, 'FFD700'),
        ('LOW', low, '90EE90'),
        ('Total Active', len(leads), '1F4E79'),
    ]

    for i, (label, count, color) in enumerate(summary_data):
        row = 5 + i
        ws3.cell(row=row, column=1, value=label)
        ws3.cell(row=row, column=2, value=count)
        ws3.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        if color in ('FF4444', '1F4E79'):
            ws3.cell(row=row, column=1).font = Font(color='FFFFFF', bold=True)

    # Save
    filepath = f'C:\\tmp\\albadi_leads_{TODAY.strftime("%Y%m%d")}.xlsx'
    wb.save(filepath)
    return filepath

# =============================================
# FETCH ALL SUBSCRIBERS
# =============================================
print(f"Fetching {len(ALL_SUBS)} leads from ManyChat...\n")
leads = []
all_leads = []  # includes skipped
skipped = 0
errors = 0

for i, sid in enumerate(ALL_SUBS):
    print(f"  [{i+1}/{len(ALL_SUBS)}] Fetching {sid}...", end=" ")
    try:
        r = requests.get(f'{BASE}/subscriber/getInfo', headers=H,
                         params={"subscriber_id": sid}, timeout=TIMEOUT)
    except requests.exceptions.Timeout:
        print("TIMEOUT")
        errors += 1
        continue
    except requests.exceptions.ConnectionError:
        print("CONNECTION ERROR")
        errors += 1
        continue

    data = r.json()

    if data.get('status') != 'success':
        print(f"ERROR: {data.get('message', 'unknown')}")
        errors += 1
        continue

    sub = data.get('data', {})
    name = sub.get('name', sid)
    tags = sub.get('tags', [])
    custom_fields = sub.get('custom_fields', [])
    phone = PHONES.get(sid, sub.get('phone', ''))

    tag_ids = [t.get('id') for t in tags]

    # Get tag name
    status_name = 'ליד_חדש'
    status_score = 0
    for tid in tag_ids:
        if tid in TAG_SCORES:
            name_t, score_t = TAG_SCORES[tid]
            if score_t > status_score:
                status_name = name_t
                status_score = score_t
    # Check for skip tags
    if any(tid == 84604877 for tid in tag_ids): status_name = 'לא_רלוונטי'
    if any(tid == 84604878 for tid in tag_ids): status_name = 'לקוח'

    notes = get_field(custom_fields, F_NOTES)
    qt_raw = get_field(custom_fields, F_QUOTE_TOTAL)
    quote_total = float(qt_raw) if qt_raw else 0
    qty_raw = get_field(custom_fields, F_QUANTITY)
    try:
        quantity = int(str(qty_raw).replace(',', '').strip()) if qty_raw else 0
    except:
        quantity = 0
    follow_up = get_field(custom_fields, F_FOLLOW_UP)
    last_contact = get_field(custom_fields, F_LAST_CONTACT)

    is_inactive = any(tid in SKIP_TAGS for tid in tag_ids)

    if is_inactive:
        print(f"SKIP ({name} - inactive)")
        skipped += 1
        all_leads.append({
            'sid': sid, 'name': name, 'phone': phone, 'tag': status_name,
            'notes': notes, 'quote_total': quote_total, 'quantity': quantity,
            'follow_up': follow_up, 'last_contact': last_contact,
            'score': 0, 'active': False,
        })
        continue

    # Calculate scores
    s_qty = quantity_score(quantity)
    s_urgency = urgency_score(follow_up)
    s_silence = silence_score(last_contact)
    total = status_score + s_qty + s_urgency + s_silence

    lead_data = {
        'sid': sid, 'name': name, 'phone': phone, 'tag': status_name,
        'notes': notes, 'quote_total': quote_total, 'quantity': quantity,
        'follow_up': follow_up, 'last_contact': last_contact,
        'score': total, 'active': True,
        'breakdown': f"status={status_score} qty={s_qty} urgency={s_urgency} silence={s_silence}",
    }
    leads.append(lead_data)
    all_leads.append(lead_data)

    print(f"OK {name} [{total}]")
    time.sleep(0.15)

print(f"\nFetched: {len(leads)} active | Skipped: {skipped} inactive | Errors: {errors}")

# Sort by score descending
leads.sort(key=lambda x: x['score'], reverse=True)

# =============================================
# SAVE SCORES TO MANYCHAT
# =============================================
print(f"\nSaving scores to ManyChat (lead_score)...")
saved = 0
for i, lead in enumerate(leads):
    print(f"  [{i+1}/{len(leads)}] {lead['name']} -> score={lead['score']}...", end=" ")
    ok = save_score_to_manychat(lead['sid'], lead['score'])
    if ok:
        saved += 1
        print("OK")
    else:
        print("FAIL")
    time.sleep(0.1)
print(f"  Saved: {saved}/{len(leads)}")

# =============================================
# PRINT DAILY CALL LIST
# =============================================
print(f"""
{'='*60}
  DAILY CALL LIST - ALBADI
  {TODAY.strftime('%Y-%m-%d')} | {len(leads)} active leads
{'='*60}
""")

for i, lead in enumerate(leads):
    name = lead['name']
    phone = lead['phone'] if lead['phone'] else '(no phone)'
    tag = lead['tag']
    score = lead['score']
    notes = lead['notes'][:80] + '...' if len(lead['notes']) > 80 else lead['notes']
    qty = f"{int(lead['quantity']):,}" if lead.get('quantity', 0) > 0 else '-'
    qt = f"{int(lead['quote_total']):,}" if lead['quote_total'] > 0 else '-'
    fu = lead['follow_up'] if lead['follow_up'] else '-'
    breakdown = lead['breakdown']

    if score >= 70: priority = "CALL NOW"
    elif score >= 50: priority = "HIGH"
    elif score >= 35: priority = "MEDIUM"
    else: priority = "LOW"

    print(f"  #{i+1} [{score}/100] {priority}")
    print(f"  {name} | {phone}")
    print(f"  Tag: {tag} | Qty: {qty} | Quote: {qt} ILS | Follow-up: {fu}")
    print(f"  Notes: {notes}")
    print(f"  ({breakdown})")
    print()

# Summary
call_now = len([l for l in leads if l['score'] >= 70])
high = len([l for l in leads if 50 <= l['score'] < 70])
medium = len([l for l in leads if 35 <= l['score'] < 50])
low = len([l for l in leads if l['score'] < 35])

print(f"""{'='*60}
  SUMMARY
  CALL NOW: {call_now} | HIGH: {high} | MEDIUM: {medium} | LOW: {low}
{'='*60}""")

# =============================================
# SAVE TO EXCEL
# =============================================
print("\nGenerating Excel...")
excel_path = save_to_excel(leads, all_leads)
print(f"  Saved: {excel_path}")

print("\nDone!")
